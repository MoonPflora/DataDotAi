import os
import csv
from openpyxl import load_workbook
from config import get_working_dir

def delete_row(json_data, current_content, row1, row2=None, row3=None, row4=None, filepath=None):
    """
    Delete one or more rows from an Excel (.xlsx) or CSV file.
    First param is filename, then up to 4 row numbers (1-based), optional filepath as last param.
    """
    # Collect rows to delete
    row_params = [row1, row2, row3, row4]
    rows_to_delete = []
    for r in row_params:
        if r is not None:
            try:
                row_num = int(r)
                if row_num < 1:
                    raise ValueError()
                rows_to_delete.append(row_num)
            except Exception:
                raise ValueError(f"Invalid row number: {r}")
    if not rows_to_delete:
        raise ValueError("At least one valid row number must be provided.")

    # Determine file path
    if filepath is None:
        working_dir = get_working_dir()
        filepath = os.path.join(working_dir, "current.xlsx")
    else:
        filepath = filepath.strip()

    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found: {filepath}")

    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".xlsx":
        wb = load_workbook(filepath)
        sheet = wb.active
        # Delete rows in descending order to avoid shifting
        for row in sorted(rows_to_delete, reverse=True):
            if row <= sheet.max_row:
                sheet.delete_rows(row)
        wb.save(filepath)

    elif ext == ".csv":
        with open(filepath, "r", newline="", encoding="utf-8") as f:
            reader = list(csv.reader(f))
        # Remove rows (1-based)
        new_rows = [row for idx, row in enumerate(reader, start=1) if idx not in rows_to_delete]
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(new_rows)

    else:
        raise ValueError(f"Unsupported file type: {ext}")

    return f"Deleted rows {rows_to_delete} from {os.path.basename(filepath)}"
