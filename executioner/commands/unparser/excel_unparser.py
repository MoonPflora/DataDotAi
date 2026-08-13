import json
import os
import openpyxl
from config import get_working_dir

def parse_cell_string(cell_str):
    """
    Parse a single parser-format string: "Sheet,row,col,\"value\""
    Returns: (sheet_name:str, row:int, col:int, value:str)
    """
    try:
        # Split only the first three commas, because value may contain commas
        parts = cell_str.split(",", 3)
        if len(parts) != 4:
            raise ValueError(f"Invalid cell string: {cell_str}")

        sheet = parts[0].strip()
        row = int(parts[1].strip())
        col = int(parts[2].strip())
        value = parts[3].strip()

        # Remove enclosing quotes and unescape
        if value.startswith('"') and value.endswith('"'):
            value = value[1:-1]
        value = value.replace('\\"', '"').replace('\\\\', '\\').replace('\\n', '\n')
        return sheet, row, col, value

    except Exception as e:
        raise ValueError(f"Error parsing cell string '{cell_str}': {e}")


def unparse_excel(parser_json, output_filename):
    """
    Convert AI-normalized parser JSON array into an Excel/XLSX file.
    - parser_json: JSON array (string or list) of cells in parser format
    - output_filename: output file name (e.g., 'names.xlsx')
    """
    # Ensure we have a list of strings
    if isinstance(parser_json, str):
        try:
            cells = json.loads(parser_json)
        except Exception as e:
            raise ValueError(f"Failed to parse JSON string: {e}")
    else:
        cells = parser_json

    if not isinstance(cells, list):
        raise ValueError("Parser input must be a JSON array of strings")

    # Organize cells by sheet
    sheets = {}
    for cell_str in cells:
        sheet, row, col, value = parse_cell_string(cell_str)
        if sheet not in sheets:
            sheets[sheet] = {}
        sheets[sheet][(row, col)] = value

    # Prepare workbook
    wb = openpyxl.Workbook()
    # Remove default sheet if unused
    if "Sheet" in wb.sheetnames and "Sheet" not in sheets:
        wb.remove(wb.active)

    # Add sheets and populate
    for sheet_name, cell_map in sheets.items():
        safe_name = sheet_name[:31] if sheet_name else "Sheet"
        ws = wb.create_sheet(title=safe_name)

        for (r, c), val in cell_map.items():
            # openpyxl is 1-based indexing
            ws.cell(row=r + 1, column=c + 1, value=val)

    # Resolve output path in working directory
    working_dir = get_working_dir()
    if not os.path.exists(working_dir):
        os.makedirs(working_dir)

    output_path = os.path.join(working_dir, output_filename)

    # Save workbook
    wb.save(output_path)

    return output_path  # optionally return path for confirmation
