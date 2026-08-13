import os
import json
import re
from openpyxl import load_workbook
from config import get_working_dir

# Path to executioner/commands/current.txt
EXECUTIONER_COMMANDS_DIR = os.path.dirname(os.path.abspath(__file__))
CURRENT_FILE = os.path.join(EXECUTIONER_COMMANDS_DIR, "current.txt")


def append(json_data: dict, *args):
    """
    Append the 'data' field from a JSON object to current.txt or other files.

    Usage:
        append(json_data, [filepath])

    json_data: dictionary containing 'data' field (string)
    filepath: optional; if omitted, append to current.txt in JSON array format
              if provided, can be .txt/.md/.csv/.log/.json/.xlsx
    """
    if "data" not in json_data:
        raise ValueError("JSON object must have a 'data' field to append.")

    # --- Unwrap JSON string to Python object ---
    raw_data = json_data["data"]
    try:
        unwrapped = json.loads(raw_data) if raw_data else []
    except json.JSONDecodeError:
        unwrapped = [raw_data]

    filepath = args[0] if len(args) > 0 else None

    # --- Append to current.txt ---
    if filepath is None:
        if not os.path.exists(CURRENT_FILE):
            current_json = []
        else:
            with open(CURRENT_FILE, "r", encoding="utf-8") as f:
                try:
                    current_json = json.load(f)
                    if not isinstance(current_json, list):
                        current_json = []
                except Exception:
                    current_json = []

        # Merge unwrapped data into current_json
        if isinstance(unwrapped, list):
            current_json.extend(unwrapped)
        else:
            current_json.append(unwrapped)

        # Write back to current.txt
        with open(CURRENT_FILE, "w", encoding="utf-8") as f:
            json.dump(current_json, f, ensure_ascii=False, indent=2)

        return f"Appended data to {os.path.basename(CURRENT_FILE)}"

    # --- Append to a specific file ---
    filepath = os.path.normpath(filepath)
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found: {filepath}")

    ext = os.path.splitext(filepath)[1].lower()

    if ext in (".txt", ".md", ".json", ".csv", ".log"):
        # For text files, write unescaped string(s)
        if isinstance(unwrapped, list):
            text_to_write = "\n".join([str(item) for item in unwrapped])
        else:
            text_to_write = str(unwrapped)
        with open(filepath, "a", encoding="utf-8") as f:
            f.write(text_to_write + "\n")
        return f"Appended text data to {os.path.basename(filepath)}"

    elif ext == ".xlsx":
        # Excel append supports sheet names
        wb = load_workbook(filepath)
        lines = unwrapped if isinstance(unwrapped, list) else [str(unwrapped)]

        for line in lines:
            # Parse sheet_name,row,col,"value"
            match = re.match(r'([^,]+),(\d+),(\d+),"(.*)"', line)
            if not match:
                continue
            sheet_name = match.group(1)
            row_idx = int(match.group(2)) + 1  # Excel 1-based
            col_idx = int(match.group(3)) + 1
            val = match.group(4)

            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(title=sheet_name)
            else:
                ws = wb[sheet_name]

            ws.cell(row=row_idx, column=col_idx).value = val

        wb.save(filepath)
        return f"Appended Excel data to {os.path.basename(filepath)}"

    else:
        raise ValueError(f"Unsupported file extension for append: {ext}")
