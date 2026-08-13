import os
import random
import string
import re
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string

# ------------------------- Configuration -------------------------
EXECUTIONER_COMMANDS_DIR = os.path.dirname(os.path.abspath(__file__))
CURRENT_FILE = os.path.join(EXECUTIONER_COMMANDS_DIR, "current.txt")
OPERATORS = ["!=", ">=", "<=", "=", "<", ">", "~"]


# ------------------------- Helper Functions -------------------------
def parse_filter_condition(condition):
    condition = condition.strip()
    for op in OPERATORS:
        if op in condition:
            parts = condition.split(op, 1)
            col = parts[0].strip()
            val = parts[1].strip()
            return col, op, val
    return None, None, condition.lower()  # free-text search


def is_number(s):
    try:
        float(s)
        return True
    except Exception:
        return False


def compare(cell_value, operator, filter_value):
    if cell_value is None:
        return False
    cell_str = str(cell_value).strip()
    filter_val = filter_value.strip()

    if operator in ("<", "<=", ">", ">="):
        if is_number(cell_str) and is_number(filter_val):
            cell_num = float(cell_str)
            filter_num = float(filter_val)
            if operator == "<":
                return cell_num < filter_num
            elif operator == "<=":
                return cell_num <= filter_num
            elif operator == ">":
                return cell_num > filter_num
            elif operator == ">=":
                return cell_num >= filter_num
        return False

    cell_lower = cell_str.lower()
    filter_lower = filter_val.lower()
    if operator in ("=", "~"):
        return filter_lower in cell_lower
    elif operator == "!=":
        return filter_lower not in cell_lower
    return False


# ------------------------- Current.txt Handling -------------------------
def parse_current_txt_lines(lines):
    metadata_lines = []
    cell_entries = []
    cell_line_re = re.compile(r'^(\d+),(\d+),"(.*)";$')
    for i, line in enumerate(lines):
        line = line.strip()
        if line == "":
            metadata_lines.append((i, line))
            continue
        m = cell_line_re.match(line)
        if m:
            r, c, val = int(m.group(1)), int(m.group(2)), m.group(3)
            cell_entries.append((r, c, val))
        else:
            metadata_lines.append((i, line))
    return metadata_lines, cell_entries


def build_table_from_cells(cell_entries):
    if not cell_entries:
        return []
    max_row = max(r for r, _, _ in cell_entries)
    max_col = max(c for _, c, _ in cell_entries)
    table = [["" for _ in range(max_col + 1)] for _ in range(max_row + 1)]
    for r, c, val in cell_entries:
        table[r][c] = val
    return table


def table_to_current_txt_lines(table, metadata_lines):
    lines = []
    meta_dict = {idx: line for idx, line in metadata_lines}
    max_meta_idx = max(meta_dict.keys()) if meta_dict else -1
    cell_lines = []
    for r, row in enumerate(table):
        for c, val in enumerate(row):
            escaped_val = val.replace('"', '""')
            cell_lines.append(f'{r},{c},"{escaped_val}";')
    final_lines = []
    cell_idx = 0
    total_lines = max(len(cell_lines), max_meta_idx + 1)
    for i in range(total_lines):
        if i in meta_dict:
            final_lines.append(meta_dict[i])
        else:
            if cell_idx < len(cell_lines):
                final_lines.append(cell_lines[cell_idx])
                cell_idx += 1
            else:
                final_lines.append("")
    while cell_idx < len(cell_lines):
        final_lines.append(cell_lines[cell_idx])
        cell_idx += 1
    return final_lines


def apply_filters_to_table(table, filters):
    if not table or len(table) < 2:
        return table
    headers = [str(h).strip() for h in table[0]]
    col_map = {h.lower(): i for i, h in enumerate(headers)}
    filtered_rows = [table[0]]
    for row in table[1:]:
        keep = True
        for cond in filters:
            col, op, val = parse_filter_condition(cond)
            if col is None:
                val_lower = val.lower()
                if not any(val_lower in (str(c).lower() if c else "") for c in row):
                    keep = False
                    break
            else:
                col_lower = col.lower()
                if col_lower not in col_map:
                    keep = False
                    break
                cell_val = row[col_map[col_lower]]
                if not compare(cell_val, op, val):
                    keep = False
                    break
        if keep:
            filtered_rows.append(row)
    return filtered_rows


def filter_current_txt(filters):
    if not os.path.exists(CURRENT_FILE):
        raise FileNotFoundError(f"{CURRENT_FILE} not found.")
    with open(CURRENT_FILE, "r", encoding="utf-8") as f:
        lines = f.readlines()
    metadata_lines, cell_entries = parse_current_txt_lines(lines)
    table = build_table_from_cells(cell_entries)
    filtered_table = apply_filters_to_table(table, filters)
    new_lines = table_to_current_txt_lines(filtered_table, metadata_lines)
    with open(CURRENT_FILE, "w", encoding="utf-8") as f:
        for line in new_lines:
            f.write(line.rstrip("\n") + "\n")
    return f"Filters [{', '.join(filters)}] applied to current.txt"


# ------------------------- Excel Filtering -------------------------
def detect_header_row_ws(ws):
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row is None:
            continue
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            headers = [str(cell).strip() if cell is not None else "" for cell in row]
            return idx, headers
    return 1, []


def filter_excel_file(filters, sheet_name, filepath):
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found: {filepath}")
    wb = load_workbook(filepath, data_only=True)
    sheets_to_process = [sheet_name] if sheet_name else wb.sheetnames
    filtered_sheets = {}
    for s_name in sheets_to_process:
        ws = wb[s_name]
        header_row_idx, headers = detect_header_row_ws(ws)
        if not headers:
            continue
        col_map = {h.lower(): i for i, h in enumerate(headers)}
        filtered_rows = [headers]
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if row is None:
                continue
            keep = True
            for cond in filters:
                col, op, val = parse_filter_condition(cond)
                if col is None:
                    val_lower = val.lower()
                    if not any(val_lower in (str(c).lower() if c else "") for c in row):
                        keep = False
                        break
                else:
                    col_lower = col.lower()
                    if col_lower not in col_map:
                        keep = False
                        break
                    cell_val = row[col_map[col_lower]]
                    if not compare(cell_val, op, val):
                        keep = False
                        break
            if keep:
                filtered_rows.append(list(row))
        filtered_sheets[s_name] = filtered_rows

    random_suffix = "".join(random.choices(string.digits, k=5))
    new_filename = f"filtered_excel#{random_suffix}.xlsx"
    new_filepath = os.path.join(EXECUTIONER_COMMANDS_DIR, new_filename)
    new_wb = Workbook()
    default_sheet = new_wb.active
    new_wb.remove(default_sheet)
    for s_name, rows in filtered_sheets.items():
        ws = new_wb.create_sheet(title=s_name)
        for r in rows:
            ws.append(r)
    new_wb.save(new_filepath)
    return new_filepath, f"Created filtered Excel file: {new_filename}"


# ------------------------- CSV Filtering -------------------------
def filter_csv_file(filters, filepath):
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"File not found: {filepath}")
    table = []
    with open(filepath, "r", encoding="utf-8-sig", newline='') as f:
        reader = csv.reader(f)
        for row in reader:
            table.append([c if c is not None else "" for c in row])

    if not table:
        raise ValueError("CSV file is empty")

    filtered_table = apply_filters_to_table(table, filters)

    random_suffix = "".join(random.choices(string.digits, k=5))
    new_filename = f"filtered_csv#{random_suffix}.csv"
    new_filepath = os.path.join(EXECUTIONER_COMMANDS_DIR, new_filename)

    with open(new_filepath, "w", encoding="utf-8-sig", newline='') as f:
        writer = csv.writer(f)
        for row in filtered_table:
            writer.writerow(row)

    return new_filepath, f"Created filtered CSV file: {new_filename}"


# ------------------------- Public API -------------------------
def filter_file(filter1=None, filter2=None, filter3=None, filter4=None, sheet_name=None, filepath=None):
    filters = [f for f in (filter1, filter2, filter3, filter4) if f]
    if not filters:
        raise ValueError("At least one filter must be provided.")

    if filepath is None:
        # Fallback to current.txt
        log_msg = filter_current_txt(filters)
        return {"filtered_file": None, "log": log_msg}

    filepath = filepath.strip()
    ext = os.path.splitext(filepath)[1].lower()

    if ext in (".xlsx", ""):
        if ext == "":
            filepath += ".xlsx"
        new_filepath, log_msg = filter_excel_file(filters, sheet_name, filepath)
    elif ext == ".csv":
        new_filepath, log_msg = filter_csv_file(filters, filepath)
    else:
        raise ValueError(f"Unsupported file type: {ext}")

    return {"filtered_file": new_filepath, "log": log_msg}
